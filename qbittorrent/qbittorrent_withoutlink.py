import qbittorrentapi
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import sys


def get_script_dir():
    """获取脚本所在目录"""
    if getattr(sys, 'frozen', False):
        script_path = sys.executable
    else:
        script_path = os.path.abspath(__file__)
    return os.path.dirname(script_path)


def to_qbittorrent_path(path):
    """将路径转换为 qBittorrent 格式（使用反斜杠）"""
    if pd.isna(path) or path is None:
        return ""
    path = str(path)
    path = path.replace("\\", "/")
    path = path.strip("/")
    return path.replace("/", "\\")


def from_qbittorrent_path(path):
    """从 qBittorrent 格式转换为标准格式（正斜杠）"""
    if not path:
        return ""
    return path.replace("\\", "/")


def get_all_parent_paths(path):
    """获取路径的所有父级路径（使用正斜杠）"""
    if not path:
        return []
    parts = path.split("/")
    paths = []
    for i in range(1, len(parts) + 1):
        paths.append("/".join(parts[:i]))
    return paths


def get_default_rule_def():
    """返回默认的规则定义模板"""
    return {
        "enabled": True,
        "mustContain": "",
        "mustNotContain": "",
        "useRegex": False,
        "episodeFilter": "",
        "smartFilter": False,
        "ignoreDays": 0,
        "affectedFeeds": [],
        "savePath": "",
        "assignedCategory": "",
        "addPaused": False,
        "previouslyMatchedEpisodes": []
    }


def show_help():
    """显示帮助信息"""
    script_name = os.path.basename(sys.argv[0])
    print("=" * 60)
    print("qBittorrent RSS 管理工具")
    print("=" * 60)
    print()
    print(f"用法: python {script_name} <命令>")
    print()
    print("可用命令:")
    print("  export    - 导出当前 qBittorrent 的 RSS 配置到 Excel")
    print("  import    - 从 Excel 导入 RSS 配置到 qBittorrent")
    print("  dry-run   - 模拟导入（预览效果，不实际修改）")
    print("  info      - 显示当前 qBittorrent 的 RSS 信息")
    print()
    print("示例:")
    print(f"  python {script_name} export")
    print(f"  python {script_name} dry-run")
    print(f"  python {script_name} import")
    print()
    script_dir = get_script_dir()
    print(f"Excel 文件位置: {os.path.join(script_dir, 'rss_manager.xlsx')}")
    print("=" * 60)


def export_rss_to_excel():
    """导出 RSS 订阅和下载设置到 Excel（带必填/选填标注）"""

    script_dir = get_script_dir()
    output_file = os.path.join(script_dir, "rss_manager.xlsx")

    conn_info = {
        "host": "localhost",
        "port": 8080,
        "username": "admin",
        "password": "adminadmin",
    }

    feeds_headers = [
        ("路径", "【必填】", 30),
        ("名称", "【选填】", 20),
        ("URL", "【条件必填】订阅时必填", 50),
        ("刷新间隔(分钟)", "【选填】", 15),
        ("文章总数", "【导出只读】", 12),
        ("未读数", "【导出只读】", 12),
        ("类型", "【必填】文件夹/订阅", 10)
    ]

    rules_headers = [
        ("规则名称", "【必填】", 20),
        ("启用", "【选填】默认:是", 8),
        ("必须包含", "【选填】", 25),
        ("必须排除", "【选填】", 25),
        ("使用正则", "【选填】默认:否", 10),
        ("剧集过滤", "【选填】", 15),
        ("智能过滤", "【选填】默认:否", 10),
        ("忽略天数", "【选填】默认:0", 10),
        ("影响订阅(路径)", "【选填】可填URL或路径", 40),
        ("保存路径", "【选填】默认qBittorrent设置", 30),
        ("分类", "【选填】", 15),
        ("添加时暂停", "【选填】默认:否", 12)
    ]

    with qbittorrentapi.Client(**conn_info) as client:
        wb = Workbook()

        # Sheet 1: RSS订阅源
        ws1 = wb.active
        ws1.title = "RSS订阅源"

        ws1.append([h[0] + h[1] for h in feeds_headers])
        ws1.append([h[0] for h in feeds_headers])

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        sub_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        sub_header_font = Font(bold=True, color="000000", size=9)

        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws1[2]:
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, (_, _, width) in enumerate(feeds_headers, 1):
            ws1.column_dimensions[get_column_letter(idx)].width = width

        items = client.rss.items.with_data
        feed_rows = []

        def process_rss_tree(data, parent_path=""):
            if not isinstance(data, dict):
                return

            for key, value in data.items():
                current_path = f"{parent_path}/{key}" if parent_path else key
                current_path = from_qbittorrent_path(current_path)

                if isinstance(value, dict):
                    if 'articles' in value:
                        articles = value.get('articles', [])
                        unread = sum(1 for a in articles if not a.get('isRead', False))
                        feed_rows.append({
                            "路径": current_path,
                            "名称": key,
                            "URL": value.get('url', ''),
                            "刷新间隔": value.get('refreshInterval', ''),
                            "文章总数": len(articles),
                            "未读数": unread,
                            "类型": "订阅"
                        })
                    else:
                        feed_rows.append({
                            "路径": current_path,
                            "名称": key,
                            "URL": "",
                            "刷新间隔": "",
                            "文章总数": "",
                            "未读数": "",
                            "类型": "文件夹"
                        })
                        process_rss_tree(value, current_path)

        process_rss_tree(items)

        for row in feed_rows:
            ws1.append([
                row["路径"],
                row["名称"],
                row["URL"],
                row["刷新间隔"],
                row["文章总数"],
                row["未读数"],
                row["类型"]
            ])

        ws1.freeze_panes = 'A3'
        ws1.auto_filter.ref = f"A2:{get_column_letter(len(feeds_headers))}{len(feed_rows) + 2}"
        ws1.row_dimensions[1].height = 30
        ws1.row_dimensions[2].height = 20

        # Sheet 2: RSS下载设置
        ws2 = wb.create_sheet(title="RSS下载设置")

        ws2.append([h[0] + h[1] for h in rules_headers])
        ws2.append([h[0] for h in rules_headers])

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws2[2]:
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, (_, _, width) in enumerate(rules_headers, 1):
            ws2.column_dimensions[get_column_letter(idx)].width = width

        rules = client.rss.rules
        for rule_name, rule in rules.items():
            affected_feeds = rule.get('affectedFeeds', [])
            affected_feeds_str = ', '.join(from_qbittorrent_path(f) for f in affected_feeds)

            ws2.append([
                rule_name,
                "是" if rule.get('enabled') else "否",
                rule.get('mustContain', ''),
                rule.get('mustNotContain', ''),
                "是" if rule.get('useRegex') else "否",
                rule.get('episodeFilter', ''),
                "是" if rule.get('smartFilter') else "否",
                rule.get('ignoreDays', 0),
                affected_feeds_str,
                rule.get('savePath', ''),
                rule.get('assignedCategory', ''),
                "是" if rule.get('addPaused') else "否"
            ])

        ws2.freeze_panes = 'A3'
        ws2.auto_filter.ref = f"A2:{get_column_letter(len(rules_headers))}{len(rules) + 2}"
        ws2.row_dimensions[1].height = 30
        ws2.row_dimensions[2].height = 20

        wb.save(output_file)
        print(f"已导出到: {output_file}")
        print(f"  - RSS订阅源: {len([r for r in feed_rows if r['类型'] == '订阅'])} 个订阅")
        print(f"  - RSS文件夹: {len([r for r in feed_rows if r['类型'] == '文件夹'])} 个文件夹")
        print(f"  - 下载规则: {len(rules)} 个")


def import_rss_from_excel(dry_run=False):
    """从 Excel 导入 RSS 订阅和下载设置"""

    script_dir = get_script_dir()
    input_file = os.path.join(script_dir, "rss_manager.xlsx")

    if not os.path.exists(input_file):
        print(f"错误: 文件不存在 - {input_file}")
        return

    conn_info = {
        "host": "localhost",
        "port": 8080,
        "username": "admin",
        "password": "adminadmin",
    }

    print(f"导入文件: {input_file}")
    if dry_run:
        print("【模拟运行模式】")
    print("-" * 40)

    try:
        df_feeds = pd.read_excel(input_file, sheet_name="RSS订阅源", header=1)
        df_rules = pd.read_excel(input_file, sheet_name="RSS下载设置", header=1)

        df_feeds.columns = [c.strip() for c in df_feeds.columns]
        df_rules.columns = [c.strip() for c in df_rules.columns]

    except Exception as e:
        print(f"读取 Excel 失败: {e}")
        return

    if "路径" in df_feeds.columns:
        df_feeds["路径"] = df_feeds["路径"].apply(
            lambda x: str(x).replace("\\", "/").strip("/") if pd.notna(x) else ""
        )

    all_folders = set()

    if "类型" in df_feeds.columns and "路径" in df_feeds.columns:
        for _, row in df_feeds[df_feeds["类型"] == "文件夹"].iterrows():
            path = row["路径"]
            if path:
                for p in get_all_parent_paths(path):
                    all_folders.add(p)

        for _, row in df_feeds[df_feeds["类型"] == "订阅"].iterrows():
            path = row["路径"]
            if path:
                parent = "/".join(path.split("/")[:-1]) if "/" in path else ""
                if parent:
                    for p in get_all_parent_paths(parent):
                        all_folders.add(p)

    url_to_path = {}
    if "URL" in df_feeds.columns and "路径" in df_feeds.columns:
        for _, row in df_feeds[df_feeds["类型"] == "订阅"].iterrows():
            url = str(row["URL"]) if pd.notna(row["URL"]) else ""
            path = str(row["路径"])
            if url and path:
                url_to_path[url] = path

    print(f"需要创建的文件夹: {len(all_folders)} 个")
    for f in sorted(all_folders):
        print(f"  - {f}")

    feed_count = len(df_feeds[df_feeds["类型"] == "订阅"]) if "类型" in df_feeds.columns else 0
    print(f"RSS 订阅: {feed_count} 个")
    print(f"规则: {len(df_rules)} 个")
    print("-" * 40)

    with qbittorrentapi.Client(**conn_info) as client:
        if all_folders:
            print("创建文件夹...")
            sorted_folders = sorted(all_folders, key=lambda x: len(x.split("/")))

            for folder_path in sorted_folders:
                qb_path = to_qbittorrent_path(folder_path)
                print(f"  {folder_path}")
                if not dry_run:
                    try:
                        client.rss.add_folder(folder_path=qb_path)
                    except qbittorrentapi.Conflict409Error:
                        print(f"    已存在")
                    except Exception as e:
                        print(f"    错误: {e}")

        if "类型" in df_feeds.columns:
            feeds = df_feeds[df_feeds["类型"] == "订阅"]
            if len(feeds) > 0:
                print("\n创建 RSS 订阅...")
                for _, row in feeds.iterrows():
                    path = row.get("路径", "")
                    url = row.get("URL", "")
                    interval = row.get("刷新间隔(分钟)")

                    if not path or not url:
                        continue

                    qb_path = to_qbittorrent_path(path)
                    print(f"  {path}")
                    if not dry_run:
                        try:
                            kwargs = {"url": url, "item_path": qb_path}
                            if pd.notna(interval) and str(interval).strip() not in ["", "nan"]:
                                kwargs["refresh_interval"] = int(float(interval))
                            client.rss.add_feed(**kwargs)
                        except qbittorrentapi.Conflict409Error:
                            print(f"    已存在")
                        except Exception as e:
                            print(f"    错误: {e}")

        if len(df_rules) > 0 and "规则名称" in df_rules.columns:
            print("\n导入 RSS 下载设置...")
            for _, row in df_rules.iterrows():
                rule_name = row["规则名称"]
                if pd.isna(rule_name) or not str(rule_name).strip():
                    continue

                affected_feeds_str = ""
                if "影响订阅(路径)" in row and pd.notna(row["影响订阅(路径)"]):
                    affected_feeds_str = str(row["影响订阅(路径)"])
                elif "影响订阅" in row and pd.notna(row["影响订阅"]):
                    affected_feeds_str = str(row["影响订阅"])

                affected_feeds = []
                for item in affected_feeds_str.split(","):
                    item = item.strip()
                    if not item:
                        continue
                    if item.startswith("http"):
                        if item in url_to_path:
                            affected_feeds.append(to_qbittorrent_path(url_to_path[item]))
                    else:
                        affected_feeds.append(to_qbittorrent_path(item))

                print(f"  {rule_name}")

                if not dry_run:
                    try:
                        rule_def = get_default_rule_def()

                        def get_val(col, default=""):
                            return str(row[col]) if col in row and pd.notna(row[col]) else default

                        def get_bool(col, default=False):
                            return row[col] == "是" if col in row and pd.notna(row[col]) else default

                        def get_int(col, default=0):
                            if col in row and pd.notna(row[col]):
                                try:
                                    return int(float(row[col]))
                                except:
                                    return default
                            return default

                        rule_def.update({
                            "enabled": get_bool("启用", True),
                            "mustContain": get_val("必须包含"),
                            "mustNotContain": get_val("必须排除"),
                            "useRegex": get_bool("使用正则"),
                            "episodeFilter": get_val("剧集过滤"),
                            "smartFilter": get_bool("智能过滤"),
                            "ignoreDays": get_int("忽略天数"),
                            "affectedFeeds": affected_feeds,
                            "savePath": get_val("保存路径"),
                            "assignedCategory": get_val("分类"),
                            "addPaused": get_bool("添加时暂停")
                        })
                        client.rss.set_rule(rule_name=rule_name, rule_def=rule_def)
                    except Exception as e:
                        print(f"    错误: {e}")

    print("\n导入完成")


def get_rss_info():
    """打印 RSS 信息到控制台"""
    conn_info = {
        "host": "localhost",
        "port": 8080,
        "username": "admin",
        "password": "adminadmin",
    }

    with qbittorrentapi.Client(**conn_info) as client:
        print(f"RSS 下载器信息 - {datetime.now().isoformat()}")
        print("-" * 40)

        items = client.rss.items.with_data
        print(f"RSS 订阅数: {len(items)}")

        def print_rss_tree(data, indent=0):
            prefix = "  " * indent
            if not isinstance(data, dict):
                return
            for key, value in data.items():
                if isinstance(value, dict):
                    if 'articles' in value:
                        articles = value.get('articles', [])
                        unread = sum(1 for a in articles if not a.get('isRead', False))
                        print(f"{prefix}[订阅] {key}: {len(articles)} 篇文章 ({unread} 未读)")
                        for article in articles[:5]:
                            status = "已读" if article.get('isRead') else "未读"
                            title = article.get('title', 'N/A')[:40]
                            print(f"{prefix}  - [{status}] {title}...")
                    else:
                        print(f"{prefix}[文件夹] {key}/")
                        print_rss_tree(value, indent + 1)

        print_rss_tree(items)

        rules = client.rss.rules
        print(f"\n自动下载规则数: {len(rules)}")
        for name, rule in rules.items():
            status = "启用" if rule.get('enabled') else "禁用"
            print(f"  {name}: {status}")
            print(f"    匹配: {rule.get('mustContain')}")
            print(f"    排除: {rule.get('mustNotContain')}")
            print(f"    影响订阅: {rule.get('affectedFeeds', [])}")
            print(f"    保存到: {rule.get('savePath') or '默认路径'}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == "export":
            export_rss_to_excel()
        elif cmd == "import":
            import_rss_from_excel(dry_run=False)
        elif cmd == "dry-run":
            import_rss_from_excel(dry_run=True)
        elif cmd == "info":
            get_rss_info()
        else:
            print(f"未知命令: {cmd}")
            show_help()
    else:
        # 无参数时显示帮助
        show_help()